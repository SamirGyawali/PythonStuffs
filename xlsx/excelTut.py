from openpyxl import workbook, load_workbook


my_workbook = load_workbook('sample.xlsx')
active_worksheet = my_workbook.active
print(my_workbook.sheetnames)